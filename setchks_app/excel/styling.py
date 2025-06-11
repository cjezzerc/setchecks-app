"""Things to do with styling of Excel output, including any conversion text that
is needed in more than one routine"""

from openpyxl.styles import Alignment, Border, Side, PatternFill, NamedStyle, Font
from openpyxl.styles.colors import Color
from openpyxl.styles import numbers


message_types_text = {
    "ISSUE": "Issue",
    "FACT": "(Info)",
    "Conditional: FACT/DEBUG": "*COND*",
}

named_styles = {}

color_fills = {
    "grey": PatternFill(patternType="solid", fgColor=Color("D9D9D9")),
    "findings_identified": PatternFill(patternType="solid", fgColor=Color("C5D9F1")),
    "apricot": PatternFill(patternType="solid", fgColor=Color("FCD5B4")),
    "user_notes": PatternFill(patternType="solid", fgColor=Color("DAEEF3")),
    "light_grey_band": PatternFill(patternType="solid", fgColor=Color("E9E9E9")),
}

border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

dashed_top_border = Border(top=Side(style="dashed"))
solid_top_border = Border(top=Side(style="thick"))
solid_bottom_border = Border(bottom=Side(style="thick"))

vsmt_style_wrap_top = NamedStyle(name="vsmt_style_wrap_top")
vsmt_style_wrap_top.alignment = Alignment(wrap_text=True, vertical="top")
vsmt_style_wrap_top.number_format = numbers.FORMAT_TEXT

vsmt_style_wrap_top_number = NamedStyle(name="vsmt_style_wrap_top_number")
vsmt_style_wrap_top_number.alignment = Alignment(wrap_text=True, vertical="top")
vsmt_style_wrap_top_number.number_format = numbers.FORMAT_NUMBER

vsmt_style_wrap_top_hyperlink = NamedStyle(name="vsmt_style_wrap_top_hyperlink")
vsmt_style_wrap_top_hyperlink.alignment = Alignment(wrap_text=True, vertical="top")
vsmt_style_wrap_top_hyperlink.number_format = numbers.FORMAT_TEXT
vsmt_style_wrap_top_hyperlink.font = Font(
    name="Calibri",
    size=11,
    bold=True,
    italic=False,
    vertAlign=None,
    underline="single",
    strike=False,
    color="FF0000EE",
)

vsmt_style_wrap_top_double_hyperlink = NamedStyle(
    name="vsmt_style_wrap_top_double_hyperlink"
)
vsmt_style_wrap_top_double_hyperlink.alignment = Alignment(
    wrap_text=True, vertical="center", horizontal="center"
)
vsmt_style_wrap_top_double_hyperlink.number_format = numbers.FORMAT_TEXT
vsmt_style_wrap_top_double_hyperlink.font = Font(
    name="Calibri",
    size=11,
    bold=True,
    italic=False,
    vertAlign=None,
    underline="double",
    strike=False,
    color="FF0000EE",
)


vsmt_style_grey_row = NamedStyle(name="vsmt_style_grey_row")
vsmt_style_grey_row.fill = color_fills["grey"]
vsmt_style_grey_row.border = border
vsmt_style_grey_row.number_format = numbers.FORMAT_TEXT

vsmt_style_Fcb = NamedStyle(name="vsmt_style_Fcb")
vsmt_style_Fcb.alignment = Alignment(vertical="bottom", horizontal="center")
vsmt_style_Fcb.number_format = numbers.FORMAT_TEXT

vsmt_style_Fcbg = NamedStyle(name="vsmt_style_Fcbg")
vsmt_style_Fcbg.alignment = Alignment(vertical="bottom", horizontal="center")
vsmt_style_Fcbg.fill = color_fills["grey"]
vsmt_style_Fcbg.border = border
vsmt_style_Fcbg.number_format = numbers.FORMAT_TEXT

vsmt_style_Tlb = NamedStyle(name="vsmt_style_Tlb")
vsmt_style_Tlb.alignment = Alignment(vertical="top", horizontal="left")
vsmt_style_Tlb.number_format = numbers.FORMAT_TEXT

vsmt_style_Tlbg = NamedStyle(name="vsmt_style_Tlbg")
vsmt_style_Tlbg.alignment = Alignment(vertical="top", horizontal="left")
vsmt_style_Tlbg.fill = color_fills["grey"]
vsmt_style_Tlbg.border = border
vsmt_style_Tlbg.number_format = numbers.FORMAT_TEXT

bold_font = Font(bold=True)
bold_red_font = Font(bold=True, color="FFFF0000")

style_name = "asterisk_on_row_overview"
named_style = NamedStyle(name=style_name)
named_styles[style_name] = named_style
named_style.fill = color_fills["grey"]
named_style.alignment = Alignment(vertical="center", horizontal="center")
named_style.number_format = numbers.FORMAT_TEXT
named_style.font = Font(
    size=14,
    bold=True,
)

style_name = "header_row"
named_style = NamedStyle(name=style_name)
named_styles[style_name] = named_style
named_style.alignment = Alignment(
    wrap_text=True, vertical="center", horizontal="center"
)
named_style.number_format = numbers.FORMAT_TEXT
named_style.font = Font(
    bold=True,
)
