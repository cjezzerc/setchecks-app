"""Functions to handle Excel input and output"""

import time
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, PatternFill
from openpyxl.styles.colors import Color
from . import (
    make_analysis_by_outcome_sheet,
    make_analysis_by_row_sheet,
    make_set_analysis_sheet,
    make_row_overview_sheet,
    make_supp_tab_sheets,
    make_chk_specific_sheets,
    make_set_info_sheet,
)


import logging

logger = logging.getLogger(__name__)

timings = {}


def generate_excel_output(
    setchks_session=None,
    excel_filename=None,
    setchks_to_include="ALL",
    all_setchks=None,
    output_OK_messages=False,
):
    """Create an excel workbook from a setchks_session object and a specified list of checks to be included in the report"""

    # raise Exception("Forcing generate excel to fail")

    time00 = time.time()

    color_fills = {
        "grey": PatternFill(patternType="solid", fgColor=Color("D9D9D9")),
        "findings_identified": PatternFill(
            patternType="solid", fgColor=Color("C5D9F1")
        ),
        "apricot": PatternFill(patternType="solid", fgColor=Color("FCD5B4")),
        "user_notes": PatternFill(patternType="solid", fgColor=Color("DAEEF3")),
        "light_grey_band": PatternFill(patternType="solid", fgColor=Color("E9E9E9")),
    }

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    setchks = setchks_session.selected_setchks
    setchks_results = setchks_session.setchks_results

    # Find which of the requested checks have actually been run (silently ignores the others)
    setchks_list_to_report = []
    for setchk_code in setchks_results.keys():
        if (setchk_code in setchks_to_include) or (setchks_to_include == "ALL"):
            setchks_list_to_report.append(setchk_code)

    wb = openpyxl.Workbook()
    for i in range(0, 8):
        ws = wb.create_sheet()

    ##################################################################
    #           Set_Info sheet                               #
    ##################################################################

    ws = wb.worksheets[0]
    ws.title = "Set_Info"

    time0 = time.time()
    make_set_info_sheet.make_set_info_sheet(
        ws=ws,
        setchks_session=setchks_session,
    )
    print(f"make_set_info_sheet took {time.time()-time0} seconds")
    timings["set info sheet"] = f"{(time.time()-time0):0.6f} s"

    ##################################################################
    #           Set_analysis sheet                               #
    ##################################################################

    ws = wb.worksheets[1]
    ws.title = "Set analyses"

    time0 = time.time()
    make_set_analysis_sheet.make_set_analysis_sheet(
        ws=ws,
        setchks_list_to_report=setchks_list_to_report,
        setchks_session=setchks_session,
        color_fills=color_fills,
        border=border,
    )
    print(f"make_set_analysis_sheet took {time.time()-time0} seconds")
    timings["set analysis sheet"] = f"{(time.time()-time0):0.6f} s"

    ##################################################################
    #           Make supp tabs sheet                                 #
    ##################################################################

    time0 = time.time()
    final_supp_tab_i_ws, supp_tabs_row_numbers_map = (
        make_supp_tab_sheets.make_supp_tab_sheets(
            wb=wb,
            first_i_ws=5,
            setchks_list_to_report=setchks_list_to_report,
            setchks_session=setchks_session,
            color_fills=color_fills,
            border=border,
        )
    )
    print(f"make_supp_tab_sheets took {time.time()-time0} seconds")
    timings["supp tab sheets"] = f"{(time.time()-time0):0.6f} s"

    ##################################################################
    #           Make chk specific sheets                             #
    ##################################################################

    time0 = time.time()
    sub_timings = {}
    final_chk_specific_i_ws = make_chk_specific_sheets.make_chk_specific_sheets(
        wb=wb,
        first_i_ws=final_supp_tab_i_ws + 1,
        setchks_list_to_report=setchks_list_to_report,
        setchks_session=setchks_session,
        color_fills=color_fills,
        border=border,
        sub_timings=sub_timings,
    )
    print(f"make_chk_specific_sheets took {time.time()-time0} seconds")
    timings["check specific sheets"] = f"{(time.time()-time0):0.6f} s"
    timings["check specific sheets subtimings"] = sub_timings

    ##################################################################
    #  "Dummy" call to create  By_Row sheet, just to get the row map #                                        #
    ##################################################################

    time0 = time.time()
    row_analysis_row_numbers_map = make_analysis_by_row_sheet.make_analysis_by_row_sheet(
        ws=None,  # this flags to routine just to make the map
        setchks_list_to_report=setchks_list_to_report,
        setchks_session=setchks_session,
        output_OK_messages=output_OK_messages,
        analysis_by_outcome_row_numbers_map=None,  # as not done the by outcome file yet
        supp_tabs_row_numbers_map=supp_tabs_row_numbers_map,
    )
    print(f"Dummy make_analysis_by_row_sheet took {time.time()-time0} seconds")
    timings["Dummy analysis by row sheet"] = f"{(time.time()-time0):0.6f} s"

    ##################################################################
    #           By_Outcome sheet                                     #
    ##################################################################

    ws = wb.worksheets[4]
    ws.title = "Grp_by_Message"

    time0 = time.time()
    analysis_by_outcome_row_numbers_map = (
        make_analysis_by_outcome_sheet.make_analysis_by_outcome_sheet(
            ws=ws,
            setchks_list_to_report=setchks_list_to_report,
            setchks_session=setchks_session,
            output_OK_messages=output_OK_messages,
            row_analysis_row_numbers_map=row_analysis_row_numbers_map,
            supp_tabs_row_numbers_map=supp_tabs_row_numbers_map,
        )
    )
    print(f"make_analysis_by_outcome_sheet took {time.time()-time0} seconds")
    timings["analysis by outcome sheet"] = f"{(time.time()-time0):0.6f} s"

    ##################################################################
    #           By_Row sheet                                         #
    ##################################################################

    ws = wb.worksheets[3]
    ws.title = "Grp_By_Row"

    time0 = time.time()
    print(analysis_by_outcome_row_numbers_map)
    row_analysis_row_numbers_map = (
        make_analysis_by_row_sheet.make_analysis_by_row_sheet(
            ws=ws,
            setchks_list_to_report=setchks_list_to_report,
            setchks_session=setchks_session,
            output_OK_messages=output_OK_messages,
            analysis_by_outcome_row_numbers_map=analysis_by_outcome_row_numbers_map,
            supp_tabs_row_numbers_map=supp_tabs_row_numbers_map,
        )
    )
    print(f"make_analysis_by_row_sheet took {time.time()-time0} seconds")
    timings["analysis by row sheet"] = f"{(time.time()-time0):0.6f} s"

    ##################################################################
    #           Row_Overview sheet                                   #
    ##################################################################

    ws = wb.worksheets[2]
    ws.title = "Row_Overview"

    time0 = time.time()
    make_row_overview_sheet.make_row_overview_sheet(
        ws=ws,
        setchks_list_to_report=setchks_list_to_report,
        setchks_session=setchks_session,
        color_fills=color_fills,
        border=border,
        row_analysis_row_numbers_map=row_analysis_row_numbers_map,
    )
    print(f"make_row_overview_sheet took {time.time()-time0} seconds")
    timings["row overview sheet"] = f"{(time.time()-time0):.6f} s"

    ##################################################################
    #          Remove vlank worksheets                               #
    ##################################################################

    for ws in wb.worksheets:
        if ws.max_row == 1 and ws.max_column == 1:
            wb.remove(ws)
        # print(f"=====> sheet {ws.title} max_row={ws.max_row} {ws.max_column}")

    ##################################################################
    #          Write workbook to file                                #
    ##################################################################

    logger.debug(f"About to save file {excel_filename}")
    time0 = time.time()
    wb.save(filename=excel_filename)
    print(f"wb.save took {time.time()-time0} seconds")
    timings["wb save"] = f"{(time.time()-time0):0.6f} s"
    print(f"Total excel file generation time =  {time.time()-time00} seconds")
    timings["total excel file generation"] = f"{(time.time()-time00):0.6f} s"

    print(timings)
    return timings

    # Aide memoire snippets
    # from openpyxl.styles import Alignment, Border, Side, PatternFill
    # from openpyxl.styles.colors import Color
    # from openpyxl.utils import get_column_letter
    #
    #     ws=wb.worksheets[0]
    # for cell in list(ws.rows)[0]:
    #     cell.alignment=Alignment(text_rotation=90, horizontal="center") #45)

    # if not cell_widths:
    #     # cell_widths=[3,3,10,3,10,25,25,20,3,8,6] + [3]*14 +[40]
    #     # cell_widths=[3,3,10,3,10,12,12,12,3,8,6] + [3]*14 +[40]
    #     cell_widths=[3,3,10,3,10,12,12,12,3,8,6] + [3]*18 +[40]
    # for i, width in enumerate(cell_widths):
    #     ws.column_dimensions[get_column_letter(i+1)].width=width

    # border = Border(left=Side(style='medium'),
    #              right=Side(style='medium'),
    #              top=Side(style='medium'),
    #              bottom=Side(style='medium'))

    # grey_fill= PatternFill(patternType='solid', fgColor=Color('D9D9D9'))

    # for row in ws.iter_rows():
    #     for cell in row:
    #         cell.border = border
    #         if cell.column_letter in ["C","E","F","G","H","Z"]: # deliberate no wrap on 4 as rare entries and useuall Sheffield and makes row too wide
    #             cell.alignment=cell.alignment.copy(wrap_text=True)
    #         if cell.column_letter in ["A","C","E","G","I","K","M","O","Q","S","U","W","Y"]: # deliberate no wrap on 4 as rare entries and useuall Sheffield and makes row too wide
    #             cell.fill=grey_fill
