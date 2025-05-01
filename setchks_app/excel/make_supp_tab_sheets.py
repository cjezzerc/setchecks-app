
from openpyxl.utils import get_column_letter
from . import styling

def make_supp_tab_sheets(
    wb=None,
    first_i_ws=None, 
    setchks_list_to_report=None,
    setchks_session=None,
    color_fills=None,
    border=None,
    ):

    supp_tabs_row_numbers_map={} # keyed by setchk code
                                 # values are tuples (ws, supp_tab_row_numbers_map) 
                                   
    i_ws=first_i_ws-1
    for setchk_code in setchks_list_to_report:
        setchk_results=setchks_session.setchks_results[setchk_code]
        if setchk_results.supp_tab_blocks is not None:
            i_ws+=1
            ws=wb.worksheets[i_ws]
            short_code=setchk_code.split("_")[0]
            ws.title=f"{short_code}_suppl"
            supp_tab_row_numbers_map=make_one_supp_tab_sheet(
                ws=ws,
                setchk_code=setchk_code,
                setchk_results=setchk_results,
                color_fills=color_fills,
                border=border,
                )
            supp_tabs_row_numbers_map[setchk_code]=(ws, supp_tab_row_numbers_map)
    return i_ws, supp_tabs_row_numbers_map

def make_one_supp_tab_sheet(
    ws=None,
    setchk_code=None,
    setchk_results=None,
    color_fills=None,
    border=None,
    ):

    supp_tab_row_numbers_map=[]    # each entry in list corresponds 1:1 to a row in data file
                                    # each such entry is a row number in supp_tab that this function constructs 
                                    # so that other sheets can link to the right row on this sheet 
    headers_output=False
    # banded_row=False
    current_ws_row=0
    first_rows_of_blocks=set()
    n_cols=-1
    for i_data_row, supp_tab_entries in enumerate(setchk_results.supp_tab_blocks):
        if supp_tab_entries not in [None, []]:
            # banded_row=not banded_row
            first_row_of_block=True
            first_rows_of_blocks.add(current_ws_row)
            for supp_tab_row in supp_tab_entries:
                
                ##############################################
                #        output header row                   #
                # (delayed until have first SuppTabRow       #
                # object to get class level attributes from) #
                if not headers_output:
                    ws.append(supp_tab_row.headers) # this is a class level attribute. 
                                                    # Just pull it the first time fine a not "None" entry
                    current_ws_row+=1
                    for i, width in enumerate(supp_tab_row.cell_widths):
                        ws.column_dimensions[get_column_letter(i+1)].width=width
                    for cell in ws[ws.max_row]:
                        cell.alignment=cell.alignment.copy(wrap_text=True, horizontal="center")
                    headers_output=True
                    ws.append([]) # for filter arrows
                    current_ws_row+=1
                    n_cols=len(supp_tab_row.cell_widths)
                #                                            #
                ##############################################
                ws.append(supp_tab_row.format_as_list())
                current_ws_row+=1

                # for cell in ws[current_ws_row]:
                #     # cell.alignment=cell.alignment.copy(wrap_text=True)
                #     # cell.border = border  
                #     # if banded_row:
                #     #     cell.style=styling.vsmt_style_Tlbg
                #     # else:
                #     #     cell.style=styling.vsmt_style_Tlb
                #     if first_row_of_block:
                #         cell.border=styling.solid_top_border
                if first_row_of_block:
                    supp_tab_row_numbers_map.append(current_ws_row)
                    first_row_of_block=False
        else:
            supp_tab_row_numbers_map.append(None)
        
    ws.freeze_panes="A3"
    
    for i_row, row in enumerate(ws.iter_rows()):
        for i_cell, cell in enumerate(row):
            strval=str(cell.value)
            if len(strval)>=16 and str(cell.value)[0:16]=='=HYPERLINK("http':
                cell.style=styling.vsmt_style_wrap_top_hyperlink
            elif len(strval)>=6 and str(cell.value)[0:6]=='=HYPER':
                cell.style=styling.vsmt_style_wrap_top_double_hyperlink
            elif i_cell in [len(row)-1, len(row)-2]: # i.e. last two columns, for eff dates
                cell.style=styling.vsmt_style_wrap_top_number # (so can filter)
            else:
                cell.style=styling.vsmt_style_wrap_top
            if i_row<=0:
                cell.font = styling.bold_font # don't use style here as destroys the 45 deg slant
                                            # (should really bring that code down to this section)
                ws.row_dimensions[i_row+1].height=50
            else:
                pass # try doing nothing to see if it makes wrap work better
                # ws.row_dimensions[i_row+1].height=18
            if (i_row) in first_rows_of_blocks:
                cell.border=styling.solid_top_border
    
    if n_cols!=-1: # this is if no non header rows have been output
        filters = ws.auto_filter
        rightmost_column=get_column_letter(n_cols)
        filters.ref = f"A2:{rightmost_column}100000" # the "current row+1" puts the filter on the row below the labels
                                                                            # (otherwise it obscures some text)  
                                                                            # the "current row+100000" makes sure define a valid region
                                                                            # but there must be a better way!!!
            
    
    return supp_tab_row_numbers_map