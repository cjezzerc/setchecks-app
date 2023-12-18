
from openpyxl.utils import get_column_letter

from openpyxl.styles import PatternFill

from openpyxl.worksheet.filters import (
    FilterColumn,
    CustomFilter,
    CustomFilters,
    DateGroupItem,
    Filters,
    )

from openpyxl.styles.colors import Color

import setchks_app.setchks.setchk_definitions

from . import styling




def make_row_overview_sheet(
    ws=None, 
    setchks_list_to_report=None,
    setchks_session=None,
    color_fills=None,
    border=None,
    row_analysis_row_numbers_map=None,
    ): 

    setchks=setchks_app.setchks.setchk_definitions.setchks
    setchks_results=setchks_session.setchks_results

    current_row=0
    current_ws_row=0

    #############################################################
    # row with the slanted setchk description column headers etc 
    # and filter column below
    #############################################################

    current_row+=1
    row_cell_contents=["Input Value Set Row"]

    for setchk_code in setchks_list_to_report:
        # row_cell_contents.append(setchk_code)
        row_cell_contents.append(setchks[setchk_code].setchk_short_name)
    
    row_cell_contents.append("Issues Identified")
    row_cell_contents.append("User Notes")

    if setchks_session.table_has_header:
        header_cell_contents=[x.string for x in setchks_session.data_as_matrix[0]]
    else:
        header_cell_contents=[f"Column {i_column}" for i_column in range(0, len(setchks_session.data_as_matrix[0]))]
    row_cell_contents+=header_cell_contents
    
    ws.append(row_cell_contents)
    current_ws_row+=1

    cell_widths=[13.33]+[3.5]*(len(setchks_list_to_report)+1)+[25]*(1+len(setchks_session.data_as_matrix[0]))
    for i, width in enumerate(cell_widths):
        ws.column_dimensions[get_column_letter(i+1)].width=width     
        ws.column_dimensions[get_column_letter(i+1)].width=width     
    ws.row_dimensions[current_row].height=136.8
    text_rotations=[0]+[45]*(len(setchks_list_to_report)+1)+[0]*(1+len(setchks_session.data_as_matrix[0]))
    filters = ws.auto_filter
    rightmost_column=get_column_letter(len(cell_widths))
    filters.ref = f"A{current_row+1}:{rightmost_column}{current_row+100000}" # the "current row+1" puts the filter on the row below the labels
                                                                        # (otherwise it obscures some text)  
                                                                        # the "current row+100000" makes sure define a valid region
                                                                        # but there must be a better way!!!
    for i_col, cell in enumerate(ws[current_row]):
        if text_rotations[i_col]==0:
            wrap_text=True
            horizontal='center'
            vertical='bottom'
        else:
            wrap_text=None
            horizontal='center' # trial and error to get these right for 45deg
            vertical='bottom' # as vertical and horizontal seem to swap meaning
        cell.alignment=cell.alignment.copy(
            text_rotation=text_rotations[i_col], 
            wrap_text=wrap_text, 
            horizontal=horizontal,
            vertical=vertical,
            )
        if cell.value=="Issues Identified":
            cell.fill=color_fills["findings_identified"]
        cell.border = border  

        col = FilterColumn(colId=0) # for column A
        col.filters = Filters()
        filters.filterColumn.append(col) # add filter to the worksheet

    current_row+=1 # to account for the filter row
    ws.append([])  # ditto
    current_ws_row+=1

    ##########################
    # row of check numbers 
    ##########################

    current_row+=1
    row_cell_contents=["CHK# (not for prod)"]

    for setchk_code in setchks_list_to_report:
        row_cell_contents.append(setchk_code[3:5])
    
    ws.append(row_cell_contents)
    current_ws_row+=1


    ################
    # the data rows 
    ################

    for i_data_row, data_row in enumerate(setchks_session.data_as_matrix[setchks_session.first_data_row:]):
        current_row+=1
        at_least_one_x=False
        x_cells=[]
        # not_blank_row_flag=not(setchks_session.marshalled_rows[i_data_row].blank_row)
        for setchk_code in setchks_list_to_report:
            setchk_results=setchks_results[setchk_code]
            if setchk_results.row_analysis!=[]:
                nothing_output_in_this_cell_yet=True # this makes sure only one thing put out per 
                                                     # row for a particular setchk
                                                     # link will be to the first one found
                                                     # (better behaviours could be implemented some day)
                for check_item in setchk_results.row_analysis[i_data_row]:
                    # if not_blank_row_flag and (check_item["Result_id"] not in [0]):
                    if nothing_output_in_this_cell_yet:
                        if check_item.outcome_level not in ["FACT","INFO","DEBUG"]:
                            row_to_link_to=row_analysis_row_numbers_map[i_data_row][setchk_code]
                            x_cells.append(f'=HYPERLINK("#By_Row!C{row_to_link_to}","x")')
                            at_least_one_x=True
                            nothing_output_in_this_cell_yet=False 
                if nothing_output_in_this_cell_yet:
                    x_cells.append("")
            else:
                x_cells.append("")
        if at_least_one_x:
            x_cells.append("*")
        else:
            x_cells.append("") 

        user_notes=[""]
        data_row_cell_contents=[x.string for x in data_row]
        row_cell_contents=[f"Row {i_data_row+setchks_session.first_data_row+1}"]
        row_cell_contents+=x_cells
        row_cell_contents+=user_notes
        row_cell_contents+=data_row_cell_contents
        ws.append(row_cell_contents)
        current_ws_row+=1

        for i_col, cell in enumerate(ws[current_row]):
            # if i_col<=len(x_cells)+1:
            #     wrap_text=False
            #     horizontal='center'
            #     vertical='bottom'
            # else:
            #     wrap_text=True
            #     horizontal='left' 
            #     vertical='bottom'
            # cell.alignment=cell.alignment.copy(
            #     wrap_text=wrap_text, 
            #     horizontal=horizontal,
            #     vertical=vertical,
            #     )
            # if i_col==len(x_cells):
            #     cell.fill=color_fills["findings_identified"]
            # if i_col==len(x_cells)+1:
            #     cell.fill=color_fills["user_notes"]
            # if at_least_one_x:
            #     cell.fill=color_fills["apricot"]
            # cell.border = border 
            if i_col<=len(x_cells)+1:
                # wrap_text=False
                # horizontal='center'
                # vertical='bottom'
                style="vsmt_style_Fcb"
            else:
                # wrap_text=True
                # horizontal='left' 
                # vertical='bottom'
                style="vsmt_style_Tlb"
            # cell.alignment=cell.alignment.copy(
            #     wrap_text=wrap_text, 
            #     horizontal=horizontal,
            #     vertical=vertical,
            #     )
            colour=""
            if i_col==len(x_cells):
                # cell.fill=color_fills["findings_identified"]
                colour="g"
            if i_col==len(x_cells)+1:
                # cell.fill=color_fills["user_notes"]
                colour="g"
            if at_least_one_x:
                # cell.fill=color_fills["apricot"]
                colour="g"
            # cell.border = border 
            cell.style=getattr(styling,style+colour)  
    
    # # crude cell with setting
    # cell_widths=[15,30,50,25,50] + [20]*10
    # for i, width in enumerate(cell_widths):
    #     ws.column_dimensions[get_column_letter(i+1)].width=width     

    # # example bit of formatting bling
    # irow=0
    # for row in ws.iter_rows():
    #     irow+=1
    #     divider_line=row[0].value=="----"
    #     for cell in row:
    #         cell.alignment=cell.alignment.copy(wrap_text=True, vertical='top')
    #         # if cell.column_letter in ["A","B","C"] or divider_line:
    #         if divider_line:
    #             cell.fill=grey_fill
    #             cell.border = border  
    #             ws.row_dimensions[irow].height = 3
