
import time
from openpyxl.utils import get_column_letter
from . import styling



def make_chk_specific_sheets(
    wb=None,
    first_i_ws=None, 
    setchks_list_to_report=None,
    setchks_session=None,
    color_fills=None,
    border=None,
    sub_timings=None,
    ):

    i_ws=first_i_ws-1
    for setchk_code in setchks_list_to_report:
        setchk_results=setchks_session.setchks_results[setchk_code]
        if setchk_results.chk_specific_sheet is not None:
            i_ws+=1
            ws=wb.worksheets[i_ws]
            ws.title=setchk_results.chk_specific_sheet.sheet_name # sf"{setchk_code}_s"
            make_one_chk_specific_sheet(
                i_ws=i_ws,
                ws=ws,
                setchk_code=setchk_code,
                setchk_results=setchk_results,
                color_fills=color_fills,
                border=border,
                sub_timings=sub_timings,
                )

def make_one_chk_specific_sheet(
    i_ws=None, # for sub_timings
    ws=None,
    setchk_code=None,
    setchk_results=None,
    color_fills=None,
    border=None,
    sub_timings=None,
    ):

    chk_specific_sheet=setchk_results.chk_specific_sheet

    current_ws_row=0
    sub_timings[f"style_{i_ws}"]=0           
    sub_timings[f"loop_{i_ws}"]=0           
    sub_timings[f"border_{i_ws}"]=0           
    for chk_specific_row in chk_specific_sheet.rows:
        # print(chk_specific_row.row_fill, chk_specific_row.row_height, chk_specific_row.cell_contents)
        ws.append(chk_specific_row.cell_contents)
        current_ws_row+=1
        if chk_specific_row.row_height is not None:
            pass
            ws.row_dimensions[current_ws_row].height = chk_specific_row.row_height
        time00=time.time()
        for cell in ws[current_ws_row]:
            # # # cell.alignment=cell.alignment.copy(wrap_text=True)
            # if chk_specific_row.row_fill is not None:
            #     cell.style=styling.vsmt_style_grey_row # only grey available as stopgap measure
            #     # cell.fill=color_fills[chk_specific_row.row_fill]
            # else:
            time0=time.time()
            cell.style=styling.vsmt_style_wrap_top 
            sub_timings[f"style_{i_ws}"]+=time.time()-time0   
            time000=time.time()        
            if chk_specific_row.is_end_of_clause:
                cell.border=styling.solid_bottom_border
            sub_timings[f"border_{i_ws}"]+=time.time()-time000   
        sub_timings[f"loop_{i_ws}"]+=time.time()-time00

    for i, width in enumerate(chk_specific_sheet.col_widths):
        ws.column_dimensions[get_column_letter(i+1)].width=width     

    ws.freeze_panes="A3"
    
    for i_row in range(0,2):
        row=ws[i_row+1]
        for i_cell, cell in enumerate(row):
            cell.style=styling.named_styles["header_row"]

    # for i_row, row in enumerate(ws.iter_rows()):
    #     # if i_row<=1:
    #     #     ws.row_dimensions[i_row+1].height = 50
    #     # else:
    #     #     pass
    #     for i_cell, cell in enumerate(row):
    #         if i_row<=1:
    #             cell.style=styling.named_styles["header_row"]
    #         else:
    #             strval=str(cell.value)
    #             if len(strval)>=16 and str(cell.value)[0:16]=='=HYPERLINK("http':
    #                 cell.style=styling.vsmt_style_wrap_top_hyperlink
    #             elif len(strval)>=6 and str(cell.value)[0:6]=='=HYPER':
    #                 cell.style=styling.vsmt_style_wrap_top_double_hyperlink
    #             else:
    #                 cell.style=styling.vsmt_style_wrap_top
    #             if (i_row+1) in top_dashed_cells and i_cell<=8:
    #                 cell.border=styling.dashed_top_border
    #             if (i_row) in divider_rows:
    #                 cell.border=styling.solid_top_border