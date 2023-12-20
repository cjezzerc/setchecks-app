
from openpyxl.utils import get_column_letter
import setchks_app.setchks.setchk_definitions
from . import styling
from .termbrowser import termbrowser_hyperlink

def make_analysis_by_outcome_sheet(
    ws=None, 
    setchks_list_to_report=None,
    setchks_session=None,
    output_OK_messages=None,
    row_analysis_row_numbers_map=None,
    supp_tabs_row_numbers_map=None,
    ): 

    setchks=setchks_app.setchks.setchk_definitions.setchks
    setchks_results=setchks_session.setchks_results

    analysis_by_outcomes_row_numbers_map=[] # each entry in list corresponds 1:1 to a row in data file
                                # each such entry is a dict
                                #        keyed by:setchk code
                                #        value=row number in analysis_by_outcome_sheet that this function constructs 
                                # structure is used so that other sheets can link to the right row on this sheet 

    # simple header row (but only if data matrix had one; need to do this better)
    current_ws_row=0
    if setchks_session.table_has_header:
        header_row_cell_contents=[x.string for x in setchks_session.data_as_matrix[0]]
        # ws.append(["Row number", "Check", "Message"] + setchks_session.data_as_matrix[0]) # ** need to create better header row
        # ws.append(["","Row specific info","Link to Grp_by_Row Tab","Row number"] + header_row_cell_contents) # ** need to create better header row
        ws.append(
            [
                "Check Number", 
                "Check Name",
                "Message Code",
                "Severity", 
                "Message",
                "Message Extension",
                "Link (R)",
                "Link (S)",
                "Input File Row Number", 
                ] + header_row_cell_contents
            ) 
        current_ws_row+=1
    ###################################################################################################
    # first loop over all check items and build a dict so can output them grouped by setchk and outcome
    ###################################################################################################
    check_items_dict={} # will be keyed by setchk_code, then by outcome_code
                        # value will be tuple of  (i_data_row, data_row, check_item,)
    for i_data_row, data_row in enumerate(setchks_session.data_as_matrix[setchks_session.first_data_row:]):
        for setchk_code in setchks_list_to_report:
            if setchk_code not in check_items_dict:
                check_items_dict[setchk_code]={}
            setchk_results=setchks_results[setchk_code]
            if setchk_results.row_analysis!=[]:
                for check_item in setchk_results.row_analysis[i_data_row]:
                    # output_OK_messages=True  # TEMPORARY while implementing
                    if output_OK_messages or check_item.outcome_level not in ["INFO","DEBUG"]:
                        outcome_code=check_item.outcome_code
                        if outcome_code not in check_items_dict[setchk_code]:
                            check_items_dict[setchk_code][outcome_code]=[]
                        check_items_dict[setchk_code][outcome_code].append((i_data_row, data_row, check_item,))


    ####################################
    # now loop over dict and make output 
    ####################################
    analysis_by_outcomes_row_numbers_map={} # keyed by outcome_code
                                            # value is dict of data_row:output_sheet_row
    divider_rows_between_checks=set()
    divider_rows_between_messages=set()
    for setchk_code in check_items_dict:
        # ws.append([setchks[setchk_code].setchk_short_name])
        if setchk_code in supp_tabs_row_numbers_map: # i.e. there is a supp tab for this check
            supp_tab_ws, supp_tab_mapping=supp_tabs_row_numbers_map[setchk_code]
            # print(f"supp_tab_mapping:{supp_tab_mapping}")
        else:
            supp_tab_ws=None  

        ws.append([
                    setchks[setchk_code].setchk_short_code, 
                    setchks[setchk_code].setchk_short_name,
                    ]
                    )
        current_ws_row+=1
        # ws.append(["----"]) 
        # current_ws_row+=1
        outcome_codes_sorted=sorted(check_items_dict[setchk_code].keys())
        for outcome_code in outcome_codes_sorted:
            analysis_by_outcomes_row_numbers_map[outcome_code]={}
            i_temp, data_row_temp, first_check_item=check_items_dict[setchk_code][outcome_code][0]
            # message=f"{outcome_code}:{first_check_item.general_message}"
            message=f"{first_check_item.general_message}"
            message_type=styling.message_types_text[first_check_item.outcome_level]
            ws.append([
                    "",
                    "",
                    outcome_code,
                    message_type,
                    message,
                    ]
                    )
            # ws.append([message])
            current_ws_row+=1
            mixed_column=setchks_session.columns_info.mixed_column
            for i_data_row, data_row, check_item in check_items_dict[setchk_code][outcome_code]:
                # data_row_cell_contents=[x.string for x in data_row]
                cid=setchks_session.marshalled_rows[i_data_row].C_Id
                cid_entered=setchks_session.marshalled_rows[i_data_row].C_Id_entered
                data_row_cell_contents=[]
                for i_col, cell_content in enumerate(data_row): 
                    if i_col==mixed_column and cid is not None and cid_entered is not None: # don't hyperlink D_Id
                                                                                            # or non-SCTID
                                                                                            # hyperlink D_Id is confusing and the implied C_Id is given nearby
                        data_row_cell_contents.append(
                            termbrowser_hyperlink(
                                sctid=cell_content.string, 
                                destination_sctid=cid,
                                )
                            )
                    else:
                        data_row_cell_contents.append(cell_content.string)
                input_file_row_number=i_data_row+setchks_session.first_data_row+1
                
                row_to_link_to=row_analysis_row_numbers_map[i_data_row][setchk_code]
                link_cell=(f'=HYPERLINK("#Grp_by_Row!H{row_to_link_to}", "R")')
                if supp_tab_ws is not None:
                    # print(f"supp_tab_mapping:{supp_tab_mapping} {i_data_row} {supp_tab_mapping[i_data_row]}")    
                    if supp_tab_mapping[i_data_row] is not None:
                        row_to_link_to=supp_tab_mapping[i_data_row]
                        # print(f"row_to_link_to {i_data_row} {row_to_link_to}")
                        supp_tab_hyperlink_cell_contents=f'=HYPERLINK("#{supp_tab_ws.title}!A{row_to_link_to}","S")'
                    else:
                        supp_tab_hyperlink_cell_contents=""
                else:
                    supp_tab_hyperlink_cell_contents=""
                row_specific_message=check_item.row_specific_message
                if row_specific_message=="None":
                    row_specific_message=""


                # ws.append([
                #     f"Row {i_data_row+setchks_session.first_data_row+1}",
                #     setchk_short_code, 
                #     setchk_short_name, 
                #     outcome_code,
                #     message_type,
                #     message,
                #     row_specific_message,
                #     hyperlink_cell_contents,
                #     supp_tab_hyperlink_cell_contents,
                #     ]
                    
                ws.append([
                    "",
                    "",
                    "",
                    "",
                    "",
                    row_specific_message,
                    link_cell,
                    supp_tab_hyperlink_cell_contents,
                    f"Row {input_file_row_number}",
                    ] + data_row_cell_contents
                    ) 
                # ws.append([
                #     "",
                #     row_specific_message,
                #     link_cell,
                #     f"Row{input_file_row_number}",
                #     ] 
                #     + data_row_cell_contents
                #     )
                current_ws_row+=1
                if i_data_row not in analysis_by_outcomes_row_numbers_map[outcome_code]:
                    analysis_by_outcomes_row_numbers_map[outcome_code][i_data_row]=[] # that this is a list is for the rare case where
                                                                                      # the same outcome code might occur more than once for the same row
                                                                                      # see comment in make_analysis_by_row_sheet.py
                analysis_by_outcomes_row_numbers_map[outcome_code][i_data_row].append(current_ws_row)
            divider_rows_between_messages.add(current_ws_row)
            # ws.append(["----"]) 
            # current_ws_row+=1
        divider_rows_between_checks.add(current_ws_row)
        # ws.append(["----"]) 
        # current_ws_row+=1
        # ws.append(["----"])
        # current_ws_row+=1 
   
    # cell_widths=[50,30,10] + [20]*10
    cell_widths=[8,30,18,8,50,30,5,5,8,25,50] + [20]*10
    for i, width in enumerate(cell_widths):
        ws.column_dimensions[get_column_letter(i+1)].width=width     

    # # example bit of formatting bling
    # for i_row, row in enumerate(ws.iter_rows()):
    #     # divider_line=row[0].value=="----"
    #     # if divider_line:
    #     #     ws.row_dimensions[irow].height = 3
    #     for i_cell, cell in enumerate(row):
    #         # cell.alignment=Alignment(wrap_text=True, vertical='top')
    #         # cell.style=vsmt_style_wrap_top
    #         # cell.alignment=cell.alignment.copy(wrap_text=True, vertical='top')
    #         # if divider_line:
    #         #     cell.style=styling.vsmt_style_grey_row
    #         #     # cell.fill=color_fills["grey"]
    #         #     # cell.border = border
    #         # else:
    #         #     cell.style=styling.vsmt_style_wrap_top 

    #         # if divider_line:
    #         #     cell.style=styling.vsmt_style_grey_row
    #         #     # cell.fill=color_fills["grey"]
    #         #     # cell.border = border
    #         else:
    #             strval=str(cell.value)
    #             if len(strval)>=16 and str(cell.value)[0:16]=='=HYPERLINK("http':
    #                 cell.style=styling.vsmt_style_wrap_top_hyperlink
    #             elif len(strval)>=6 and str(cell.value)[0:6]=='=HYPER':
    #                 cell.style=styling.vsmt_style_wrap_top_double_hyperlink
    #             else:
    #                 cell.style=styling.vsmt_style_wrap_top

        ws.freeze_panes="J2"
        
    for i_row, row in enumerate(ws.iter_rows()):
        if i_row==0:
            ws.row_dimensions[i_row+1].height = 50
        else:
            pass
        for i_cell, cell in enumerate(row):
            if i_row==0:
                cell.style=styling.named_styles["header_row"]
            else:
                strval=str(cell.value)
                if len(strval)>=16 and str(cell.value)[0:16]=='=HYPERLINK("http':
                    cell.style=styling.vsmt_style_wrap_top_hyperlink
                elif len(strval)>=6 and str(cell.value)[0:6]=='=HYPER':
                    cell.style=styling.vsmt_style_wrap_top_double_hyperlink
                else:
                    cell.style=styling.vsmt_style_wrap_top
                if (i_row) in divider_rows_between_messages: # and i_cell<=8:
                    cell.border=styling.dashed_top_border
                if (i_row) in divider_rows_between_checks:
                    cell.border=styling.solid_top_border

    # ws['A1'].hyperlink="#By_Row!C5"
    return analysis_by_outcomes_row_numbers_map
