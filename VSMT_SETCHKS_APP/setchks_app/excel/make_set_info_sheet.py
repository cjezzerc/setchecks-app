

from openpyxl.utils import get_column_letter
import setchks_app.setchks.setchk_definitions
from . import styling
from flask import current_app


def make_set_info_sheet(
    ws=None, 
    setchks_session=None,
    ):

    ws.title='Set_Info'

    ws.append([
        "Value Set Name",
        setchks_session.vs_name,
        ])
    
    ws.append([
        "Value Set Purpose",
        setchks_session.vs_purpose,
        ])
    
    ws.append([
        "Date Checks Run",
        setchks_session.time_started_processing,
        ])
    
    ws.append([
        "Input File Name",
        setchks_session.filename,
        ])
    
    ws.append([
        "Excel Output File Name",
        setchks_session.excel_filename.split("/")[-1],
        ])
    
    context_nicer_words={
        "ENTRY_PRIMARY":"Data entry value set for Primary Care purposes",
        "ENTRY_OTHER":"Data entry value set for non-Primary Care purposes",
        "EXTRACT":"Data extraction value set",
        }
    
    ws.append([
        "Context",
        context_nicer_words[setchks_session.data_entry_extract_type],
        ])

    mode_nicer_words={
        "SINGLE_SCT_VERSION":"Set checks (Single SNOMED CT Version)",
        "DUAL_SCT_VERSIONS":"Assess changes between two releases (Dual SNOMED CT Versions)",
        }

    ws.append([
        "Mode",
        mode_nicer_words[setchks_session.sct_version_mode],
        ])

    if setchks_session.sct_version_mode=="SINGLE_SCT_VERSION":
        ws.append([
            "SNOMED CT Version",
            f"UK Monolith Edition {setchks_session.sct_version.name_for_dropdown}",
        ])
    else:
        ws.append([
            "Earlier SNOMED CT Version",
            f"UK Monolith Edition {setchks_session.sct_version.name_for_dropdown}",
        ])
        ws.append([
            "Later SNOMED CT Version",
            f"{setchks_session.sct_version_b.name_for_dropdown}",
        ])


    ws.append([
        "User email",
        setchks_session.email,
        ])

    ws.append([
        "Run identifier",
        f"{setchks_session.uuid}:{setchks_session.time_started_processing}",
        ])

    ws.append([
        "Software Version",
        "TBI"
        # current_app.config["VERSION"]
        ])
    
    cell_widths=[30,80]
    for i, width in enumerate(cell_widths):
        ws.column_dimensions[get_column_letter(i+1)].width=width     

    
    for i_row, row in enumerate(ws.iter_rows()):
        # if i_row==0:
        #     ws.row_dimensions[i_row+1].height = 50
        # else:
        #     pass
        for i_cell, cell in enumerate(row):
            cell.style=styling.vsmt_style_wrap_top
            if i_cell==0:
                cell.font=styling.bold_font
            