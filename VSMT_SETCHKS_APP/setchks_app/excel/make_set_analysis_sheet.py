

from openpyxl.utils import get_column_letter
import setchks_app.setchks.setchk_definitions
from . import styling


def make_set_analysis_sheet(
    ws=None, 
    setchks_list_to_report=None,
    setchks_session=None,
    color_fills=None,
    border=None,
    ):

    setchks=setchks_app.setchks.setchk_definitions.setchks
    setchks_results=setchks_session.setchks_results

    ws.title='Set_Analysis'
    current_ws_row=0
    # simple header row
    ws.append(
        [
        "Check Code", 
        "Check Name",
        "Message Code",
        "Severity",
        "Message",
        "Count",
        "Value",
        ]
        )
    current_ws_row+=1

    end_of_block_rows=set()
    
    # for setchk_code in setchks_list_to_report:
    for setchk_code in [x.setchk_code for x in setchks_session.selected_setchks]:
        setchk_short_name=setchks[setchk_code].setchk_short_name
        setchk_short_code=setchks[setchk_code].setchk_short_code
        if (setchk_code in setchks_session.setchks_run_status):  # test needed in case fail gatekeeper
            if setchks_session.setchks_run_status[setchk_code]=="failed":
                message_code=setchk_code.split("_")[0]+"-OUT-FAIL"
                severity="RED"
                message="Check failed. Please report to software developers."
                count=""
                value=""
                ws.append(
                    [
                    setchk_short_code, 
                    setchk_short_name, 
                    message_code,
                    severity,
                    message,
                    count,
                    value 
                    ]
                    )
                current_ws_row+=1
            else:
                for message in setchks_results[setchk_code].set_analysis["Messages"]:
                    ws.append([setchk_short_name, message])
                    current_ws_row+=1
                    # cell=ws[ws.max_row][0]
                    # cell=ws[ws.max_row][1]
                for set_level_table_row in setchks_results[setchk_code].set_level_table_rows:
                    if set_level_table_row.simple_message is not None:
                        f=set_level_table_row.simple_message.split()
                        severity=f[0][1:-1]
                        message=" ".join(f[1:])
                        message_code=set_level_table_row.outcome_code
                        count=""
                        value=""
                    else:
                        severity=""
                        message=""
                        message_code=set_level_table_row.outcome_code
                        count=set_level_table_row.descriptor
                        value=set_level_table_row.value
                    ws.append(
                        [
                        setchk_short_code, 
                        setchk_short_name, 
                        message_code,
                        severity,
                        message,
                        count,
                        value 
                        ]
                        )
                    current_ws_row+=1
        # ws.append(["----"]) 
        end_of_block_rows.add(current_ws_row)
    cell_widths=[8,32,15,8,65,65,10]
    for i, width in enumerate(cell_widths):
        ws.column_dimensions[get_column_letter(i+1)].width=width     

    # for i_row, row in enumerate(ws.iter_rows()):
    #     divider_line=row[0].value=="----"
    #     for i_cell, cell in enumerate(row):
    #         # cell.alignment=cell.alignment.copy(wrap_text=True, vertical='top')
    #         # # if cell.column_letter in ["A","B","C"] or divider_line:
    #         # if divider_line:
    #         #     cell.fill=color_fills["grey"]
    #         #     cell.border = border  
    #         #     ws.row_dimensions[irow].height = 3

    ws.freeze_panes="A6"
    
    for i_row, row in enumerate(ws.iter_rows()):
        if i_row==0:
            ws.row_dimensions[i_row+1].height = 50
        else:
            pass
        for i_cell, cell in enumerate(row):
            if i_row in [1,2,3,4] and i_cell in [0,1,2,3]:
                cell.value="" # crude way to use first four lines from results of CHK20
                                # as file level stats 
            if i_row==0:
                cell.style=styling.named_styles["header_row"]
            else:
                cell.style=styling.vsmt_style_wrap_top
            if (i_row+1) in end_of_block_rows:
                cell.border=styling.solid_bottom_border