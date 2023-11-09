"""Functions to handle Excel input and output"""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, PatternFill
from openpyxl.styles.colors import Color
from . import (
    make_analysis_by_outcome_sheet, 
    make_analysis_by_row_sheet, 
    make_set_analysis_sheet, 
    make_row_overview_sheet, 
    make_supp_tab_sheets
    )

def generate_excel_output(setchks_session=None, excel_filename=None, setchks_to_include="ALL", all_setchks=None, output_OK_messages=False):
    """Create an excel workbook from a setchks_session object and a specified list of checks to be included in the report"""
    
    color_fills={
        "grey": PatternFill(patternType='solid', fgColor=Color('D9D9D9')),
        "findings_identified": PatternFill(patternType='solid', fgColor=Color('C5D9F1')),
        "apricot": PatternFill(patternType='solid', fgColor=Color('FCD5B4')),
        "user_notes": PatternFill(patternType='solid', fgColor=Color('DAEEF3')),
        "light_grey_band": PatternFill(patternType='solid', fgColor=Color('E9E9E9')),
    }

    border = Border(left=Side(style='thin'), 
                 right=Side(style='thin'), 
                 top=Side(style='thin'), 
                 bottom=Side(style='thin'))
    

    setchks=setchks_session.available_setchks
    setchks_results=setchks_session.setchks_results
    
    #Find which of the requested checks have actually been run (silently ignores the others)
    setchks_list_to_report=[]
    for setchk_code in setchks_results.keys():
        if (setchk_code in setchks_to_include) or (setchks_to_include=="ALL"):
            setchks_list_to_report.append(setchk_code)

    wb=openpyxl.Workbook()
    for i in range(0,5):
        ws=wb.create_sheet()

    ##################################################################
    #           Set_analysis sheet                               #     
    ##################################################################

    ws=wb.worksheets[3]
    ws.title='Set analyses'

    make_set_analysis_sheet.make_set_analysis_sheet(
        ws=ws, 
        setchks_list_to_report=setchks_list_to_report,
        setchks_session=setchks_session,
        color_fills=color_fills,
        border=border,
        )

    ##################################################################
    #           Make supp tabs sheet                                 #     
    ##################################################################

    supp_tabs_row_numbers_map=make_supp_tab_sheets.make_supp_tab_sheets(
        wb=wb,
        first_i_ws=4, 
        setchks_list_to_report=setchks_list_to_report,
        setchks_session=setchks_session,
        color_fills=color_fills,
        border=border,
        )
    print(supp_tabs_row_numbers_map)

    ##################################################################
    #           By_Outcome sheet                                     #     
    ##################################################################

    ws=wb.worksheets[2]
    ws.title='By_Outcome'

    analysis_by_outcome_row_numbers_map=make_analysis_by_outcome_sheet.make_analysis_by_outcome_sheet(
        ws=ws, 
        setchks_list_to_report=setchks_list_to_report,
        setchks_session=setchks_session,
        color_fills=color_fills,
        border=border,
        output_OK_messages=output_OK_messages,
        )
    
    ##################################################################
    #           By_Row sheet                                         #     
    ##################################################################

    ws=wb.worksheets[1]
    ws.title='By_Row'

    row_analysis_row_numbers_map=make_analysis_by_row_sheet.make_analysis_by_row_sheet(
        ws=ws, 
        setchks_list_to_report=setchks_list_to_report,
        setchks_session=setchks_session,
        color_fills=color_fills,
        border=border,
        output_OK_messages=output_OK_messages,
        analysis_by_outcome_row_numbers_map=analysis_by_outcome_row_numbers_map,
        supp_tabs_row_numbers_map=supp_tabs_row_numbers_map,
        )

    ##################################################################
    #           Row_Overview sheet                                   #     
    ##################################################################

    ws=wb.worksheets[0]
    ws.title='Row_Overview'

    make_row_overview_sheet.make_row_overview_sheet(
        ws=ws, 
        setchks_list_to_report=setchks_list_to_report,
        setchks_session=setchks_session,
        color_fills=color_fills,
        border=border,
        row_analysis_row_numbers_map=row_analysis_row_numbers_map,
        )
    

    
    ##################################################################
    #          Write workbook to file                                #     
    ##################################################################

    wb.save(filename=excel_filename)

        # Aide memoire snippets
        # from openpyxl.styles import Alignment, Border, Side, PatternFill
        # from openpyxl.styles.colors import Color
        # from openpyxl.utils import get_column_letter
        #
        #     ws=wb.worksheets[0]
        # for cell in list(ws.rows)[0]:
        #     cell.alignment=Alignment(text_rotation=90, horizontal="center") #45)
        
        # if not cell_widths:
        #     # cell_widths=[3,3,10,3,10,25,25,20,3,8,6] + [3]*14 +[40]
        #     # cell_widths=[3,3,10,3,10,12,12,12,3,8,6] + [3]*14 +[40]
        #     cell_widths=[3,3,10,3,10,12,12,12,3,8,6] + [3]*18 +[40]
        # for i, width in enumerate(cell_widths):
        #     ws.column_dimensions[get_column_letter(i+1)].width=width     
        
        # border = Border(left=Side(style='medium'), 
        #              right=Side(style='medium'), 
        #              top=Side(style='medium'), 
        #              bottom=Side(style='medium'))

        # grey_fill= PatternFill(patternType='solid', fgColor=Color('D9D9D9'))

        # for row in ws.iter_rows():
        #     for cell in row:
        #         cell.border = border  
        #         if cell.column_letter in ["C","E","F","G","H","Z"]: # deliberate no wrap on 4 as rare entries and useuall Sheffield and makes row too wide
        #             cell.alignment=cell.alignment.copy(wrap_text=True)
        #         if cell.column_letter in ["A","C","E","G","I","K","M","O","Q","S","U","W","Y"]: # deliberate no wrap on 4 as rare entries and useuall Sheffield and makes row too wide
        #             cell.fill=grey_fill