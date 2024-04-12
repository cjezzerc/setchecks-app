import time

from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Alignment
import setchks_app.setchks.setchk_definitions
from . import styling
from .termbrowser import termbrowser_hyperlink
 
def make_analysis_by_row_sheet(
    ws=None, 
    setchks_list_to_report=None,
    setchks_session=None,
    output_OK_messages=None,
    analysis_by_outcome_row_numbers_map=None,
    supp_tabs_row_numbers_map=None,
    ):

    if ws is None: # if no ws passed in then the routine is just precalculating the row map
        # as there is a circular dependendency of the analysis_by_row and analysis_by_outcome sheets 
        creating_ws=False # use this to stub out most ws operations
        ws=[] # but with ws as a regular list the ws.append lines do not need to be stubbed out
    else:
        creating_ws=True

    # vsmt_style_wrap_top=NamedStyle(name="vsmt_style_wrap_top")
    # vsmt_style_wrap_top.alignment=Alignment(wrap_text=True, vertical='top')

    # vsmt_style_grey_row=NamedStyle(name="vsmt_style_grey_row")
    # vsmt_style_grey_row.fill=color_fills["grey"]
    # vsmt_style_grey_row.border=border

    setchks=setchks_app.setchks.setchk_definitions.setchks
    setchks_results=setchks_session.setchks_results
    ci=setchks_session.columns_info

    row_analysis_row_numbers_map=[] # each entry in list corresponds 1:1 to a row in data file
                                # each such entry is a dict
                                #        keyed by:setchk code
                                #        value=row number in row_analysis_sheet that this function constructs 
                                # structure is used so that other sheets can link to the right row on this sheet 

    # simple header row (but only if data matrix had one; need to do this better)
    current_ws_row=0
    if setchks_session.table_has_header:
        header_row_cell_contents=[x.string for x in setchks_session.data_as_matrix[0]]
        # ws.append(["Row number", "Check", "Message"] + setchks_session.data_as_matrix[0]) # ** need to create better header row
        identifier_term_separator_headers=[]
        identifier_header=header_row_cell_contents[ci.mixed_column]
        identifier_term_separator_headers.append(f'Identifier ("{identifier_header}")')
        if ci.have_dterm_column:
            dterm_header=header_row_cell_contents[ci.dterm_column]
            identifier_term_separator_headers.append(f'Term ("{dterm_header}")')
        identifier_term_separator_headers.append("")
        ws.append(
            [
                "Input File Row Number", 
                "Check Number", 
                "Check Name",
                "Message Code",
                "Severity", 
                "Message",
                "Message Extension",
                "Link to Grp by Msg",
                "Link to Suppl Tab",
                ] + identifier_term_separator_headers + header_row_cell_contents
            ) 
        current_ws_row+=1
        ws.append([]) # for filter arrows
        current_ws_row+=1
        
    ws_append_time=0
    time_bigloop0=time.time()
    n_loops=0
    total_dt_inner_loop=0
    max_dt_inner_loop=0
    min_dt_inner_loop=9999999
    mixed_column=setchks_session.columns_info.mixed_column
    previous_setchk_short_code=None
    divider_rows=set()
    top_dashed_cells=set()

    for i_data_row, data_row in enumerate(setchks_session.data_as_matrix[setchks_session.first_data_row:]):
        # data_row_cell_contents=[x.string for x in data_row]
        cid=setchks_session.marshalled_rows[i_data_row].C_Id
        cid_entered=setchks_session.marshalled_rows[i_data_row].C_Id_entered
        data_row_cell_contents=[]
        for i_col, cell_content in enumerate(data_row): 
            if i_col==mixed_column and cid is not None and cid_entered is not None: # don't hyperlink D_Id
                                                                                    # or non-SCTID
                                                                                    # hyperlink D_Id is confusing and the implied C_Id is given nearby
                data_row_cell_contents.append(
                    termbrowser_hyperlink(
                        sctid=cell_content.string, 
                        destination_sctid=cid,
                        )
                    )
            else:
                data_row_cell_contents.append(cell_content.string)

        
        something_was_output=False
        row_analysis_row_numbers_map.append({})
        current_row_map=row_analysis_row_numbers_map[-1] 
        for setchk_code in setchks_list_to_report:
            setchk_short_name=setchks[setchk_code].setchk_short_name
            setchk_short_code=setchks[setchk_code].setchk_short_code
            setchk_results=setchks_results[setchk_code]
            if setchk_results.row_analysis!=[]:
                if setchk_code in supp_tabs_row_numbers_map: # i.e. there is a supp tab for this check
                    supp_tab_ws, supp_tab_mapping=supp_tabs_row_numbers_map[setchk_code]
                    # print(f"supp_tab_mapping:{supp_tab_mapping}")
                else:
                    supp_tab_ws=None        
                # data_row_cell_contents=[x.string for x in data_row]
                # ws.append([i_data_row+setchks_session.first_data_row+1, setchk_short_name, setchk_results.row_analysis[i_data_row]["Message"]]+data_row_cell_contents)
                outcome_codes_count={} # this is used to make sure that in the case where the same outcome_code
                                    # can occur more than once for the same row (e.g. where checking for unreccomended tl-hierarchies
                                    # and the concept is in more than one tl-hierarchy then) then the hyperlink fgoes to the
                                    # correct row of the "by outcome" table
                                    # this has been implemented but not thoroughly tested yet! 
                for check_item in setchk_results.row_analysis[i_data_row]:
                    if output_OK_messages or check_item.outcome_level not in ["INFO","DEBUG"]:
                        time_inner_0=time.time()
                        outcome_code=check_item.outcome_code
                        
                        if outcome_code not in outcome_codes_count:
                            outcome_codes_count[outcome_code]=0
                        else:
                            outcome_codes_count[outcome_code]+=1
                        
                        if creating_ws:
                            row_to_link_to=analysis_by_outcome_row_numbers_map[outcome_code][i_data_row][outcome_codes_count[outcome_code]]
                        else:
                            row_to_link_to=None
                        
                        # message=f"{outcome_code}:{check_item.general_message}" 
                        message_type=styling.message_types_text[check_item.outcome_level]
                        message=check_item.general_message
                        row_specific_message=check_item.row_specific_message
                        if row_specific_message == "None":
                            row_specific_message=""
                        hyperlink_cell_contents=f'=HYPERLINK("#Grp_by_Message!G{row_to_link_to}","M")'
                        # print(f"i_data_row: {i_data_row} supp_tab_ws: {supp_tab_ws}")
                        if supp_tab_ws is not None:
                            # print(f"supp_tab_mapping:{supp_tab_mapping} {i_data_row} {supp_tab_mapping[i_data_row]}")    
                            if supp_tab_mapping[i_data_row] is not None:
                                row_to_link_to=supp_tab_mapping[i_data_row]
                                # print(f"row_to_link_to {i_data_row} {row_to_link_to}")
                                supp_tab_hyperlink_cell_contents=f'=HYPERLINK("#{supp_tab_ws.title}!A{row_to_link_to}","S")'
                            else:
                                supp_tab_hyperlink_cell_contents=""
                        else:
                            supp_tab_hyperlink_cell_contents=""

                        # print(f"MESSAGE_CCELL_CONTENTS:{message_cell_contents}")
                        # print(len(message_cell_contents))
                        ws_row_contents=[
                            f"Row {i_data_row+setchks_session.first_data_row+1}",
                            setchk_short_code, 
                            setchk_short_name, 
                            outcome_code,
                            message_type,
                            message,
                            row_specific_message,
                            hyperlink_cell_contents,
                            supp_tab_hyperlink_cell_contents,
                            ] 
                        # if not something_was_output: # only add the file data for the first outcome line
                        #     ws_row_contents+=data_row_cell_contents

                        identifier_term_separator_data=[]
                        identifier_data=data_row_cell_contents[ci.mixed_column]
                        identifier_term_separator_data.append(identifier_data)
                        if ci.have_dterm_column:
                            dterm_data=data_row_cell_contents[ci.dterm_column]
                            identifier_term_separator_data.append(dterm_data)
                        identifier_term_separator_data.append("Input file -->")
                        ws_row_contents+=identifier_term_separator_data

                        ws_row_contents+=data_row_cell_contents # reverted to always showing as filtering
                                                                # can lead to weird misundertstandings
                        time0=time.time()
                        ws.append(ws_row_contents)
                        current_ws_row+=1
                        dt=time.time()-time0
                        ws_append_time+=dt
                        if (setchk_short_code!=previous_setchk_short_code) and something_was_output:
                            top_dashed_cells.add(current_ws_row)
                        previous_setchk_short_code=setchk_short_code
                        n_loops+=1
                        # current_row_map[setchk_code]=ws.max_row
                        current_row_map[setchk_code]=current_ws_row
                        something_was_output=True
                        dt_inner_loop=time.time()-time_inner_0
                        total_dt_inner_loop+=dt_inner_loop
                        max_dt_inner_loop=max(max_dt_inner_loop, dt_inner_loop)
                        min_dt_inner_loop=min(min_dt_inner_loop, dt_inner_loop)
                        if dt_inner_loop>0.003:
                            print(setchk_code,i_data_row, dt_inner_loop)
        if something_was_output:
            # ws.append(["----"]) 
            divider_rows.add(current_ws_row)
            # current_ws_row+=1
    print(f"Big loop time: {time.time()-time_bigloop0}")
    print(f"ws_append_time: {ws_append_time}")
    print(f"n_loops: {n_loops}")
    print(f"total_dt_inner_loop{total_dt_inner_loop}")
    print(f"max_dt_inner_loop{max_dt_inner_loop}")
    print(f"min_dt_inner_loop{min_dt_inner_loop}")
    
    print(f"top_dashed_cells={top_dashed_cells}")
    if creating_ws:    
        # cell_widths=[8,8,30,18,8,50,30,7,7,25,50] + [20]*10
        cell_widths=[8,8,20,15,8,50,30,7,7,20]
        if ci.have_dterm_column:
            cell_widths.append(30)
        cell_widths.append(12)
        cell_widths+=[20]*ci.ncols
        for i, width in enumerate(cell_widths):
            ws.column_dimensions[get_column_letter(i+1)].width=width     

        # ncols_to_freeze=12
        # if ci.have_dterm_column:
        #     ncols_to_freeze+=1
        # # ws.freeze_panes="J3"
        # ws.freeze_panes=get_column_letter(ncols_to_freeze)+"2"
        
        for i_row, row in enumerate(ws.iter_rows()):
            if i_row==0:
                ws.row_dimensions[i_row+1].height = 50
            else:
                pass
            for i_cell, cell in enumerate(row):
                if i_row==0:
                    cell.style=styling.named_styles["header_row"]
                else:
                    strval=str(cell.value)
                    if len(strval)>=16 and str(cell.value)[0:16]=='=HYPERLINK("http':
                        cell.style=styling.vsmt_style_wrap_top_hyperlink
                    elif len(strval)>=6 and str(cell.value)[0:6]=='=HYPER':
                        cell.style=styling.vsmt_style_wrap_top_double_hyperlink
                    else:
                        cell.style=styling.vsmt_style_wrap_top
                    if (i_row+1) in top_dashed_cells and i_cell<=8:
                        cell.border=styling.dashed_top_border
                    if (i_row) in divider_rows:
                        cell.border=styling.solid_top_border
        filters = ws.auto_filter
        rightmost_column=get_column_letter(len(cell_widths))
        filters.ref = f"A2:{rightmost_column}100000" # the "current row+1" puts the filter on the row below the labels
                                                                            # (otherwise it obscures some text)  
                                                                            # the "current row+100000" makes sure define a valid region
                                                                            # but there must be a better way!!!
            

    return row_analysis_row_numbers_map

  