"""Functions to handle Excel input and output"""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, PatternFill
from openpyxl.styles.colors import Color

def generate_excel_output(setchks_session=None, excel_filename=None, setchks_to_include="ALL", all_setchks=None, output_OK_messages=False):
    """Create an excel workbook from a setchks_session object and a specified list of checks to be included in the report"""
    
    grey_fill= PatternFill(patternType='solid', fgColor=Color('D9D9D9'))
    border = Border(left=Side(style='thin'), 
                 right=Side(style='thin'), 
                 top=Side(style='thin'), 
                 bottom=Side(style='thin'))
    
    
    setchks=setchks_session.available_setchks
    setchks_results=setchks_session.setchks_results
    
    #Find which of the requested checks have actually been run (silently ignores the others)
    setchks_list_to_report=[]
    for setchk_code in setchks_results.keys():
        if (setchk_code in setchks_to_include) or (setchks_to_include=="ALL"):
            setchks_list_to_report.append(setchk_code)

    ##################################################################
    ##################################################################
    ##################################################################
    #           Sheet 1: Set analyses                                #     
    ##################################################################
    ##################################################################
    ##################################################################

    wb=openpyxl.Workbook()
    ws=wb.active
    ws.title='Set analyses'

    # simple header row
    ws.append(["Check", "Message"])

    # Add the set level messages from each check on first sheet
    for setchk_code in setchks_list_to_report:
        setchk_short_name=setchks[setchk_code].setchk_short_name
        for message in setchks_results[setchk_code].set_analysis["Messages"]:
            ws.append([setchk_short_name, message])
            cell=ws[ws.max_row][0]
            cell=ws[ws.max_row][1]
        ws.append(["----"]) 
    # crude cell with setting
    cell_widths=[30,100]
    for i, width in enumerate(cell_widths):
        ws.column_dimensions[get_column_letter(i+1)].width=width     

    for row in ws.iter_rows():
        divider_line=row[0].value=="----"
        for cell in row:
            cell.alignment=cell.alignment.copy(wrap_text=True, vertical='top')
            # if cell.column_letter in ["A","B","C"] or divider_line:
            if divider_line:
                cell.fill=grey_fill
                cell.border = border  

    ##################################################################
    ##################################################################
    ##################################################################
    #           Sheet 2: Row analyses                                #     
    ##################################################################
    ##################################################################
    ##################################################################

    # add sheet with row by row analysis
    ws=wb.create_sheet(title="Row analyses")

    # simple header row (but only if data matrix had one; need to do this better)
    if setchks_session.table_has_header:
        header_row_cell_contents=[x.string for x in setchks_session.data_as_matrix[0]]
        # ws.append(["Row number", "Check", "Message"] + setchks_session.data_as_matrix[0]) # ** need to create better header row
        ws.append(["Row number", "Check", "Message"] + header_row_cell_contents) # ** need to create better header row

    for i_data_row, data_row in enumerate(setchks_session.data_as_matrix[setchks_session.first_data_row:]):
        something_was_output=False
        for setchk_code in setchks_list_to_report:
            setchk_short_name=setchks[setchk_code].setchk_short_name
            setchk_results=setchks_results[setchk_code]
            data_row_cell_contents=[x.string for x in data_row]
            # ws.append([i_data_row+setchks_session.first_data_row+1, setchk_short_name, setchk_results.row_analysis[i_data_row]["Message"]]+data_row_cell_contents)
            for check_item in setchk_results.row_analysis[i_data_row]:
                if output_OK_messages or (check_item["Result_id"]!=0):
                    # message="%s : %s " % (check_item["Result_id"], check_item["Message"]) 
                    message="%s" % (check_item["Message"]) 
                    ws.append([i_data_row+setchks_session.first_data_row+1, setchk_short_name, message]+data_row_cell_contents)
                    something_was_output=True
        if something_was_output:
            ws.append(["----"]) 
    
    # crude cell with setting
    cell_widths=[15,30,50,25,50] + [20]*10
    for i, width in enumerate(cell_widths):
        ws.column_dimensions[get_column_letter(i+1)].width=width     

    # example bit of formatting bling
    irow=0
    for row in ws.iter_rows():
        divider_line=row[0].value=="----"
        irow+=1
        for cell in row:
            cell.alignment=cell.alignment.copy(wrap_text=True, vertical='top')
            # if cell.column_letter in ["A","B","C"] or divider_line:
            if divider_line:
                cell.fill=grey_fill
                cell.border = border  
                ws.row_dimensions[irow].height = 3
    wb.save(filename=excel_filename)

        # Aide memoire snippets
        # from openpyxl.styles import Alignment, Border, Side, PatternFill
        # from openpyxl.styles.colors import Color
        # from openpyxl.utils import get_column_letter
        #
        #     ws=wb.worksheets[0]
        # for cell in list(ws.rows)[0]:
        #     cell.alignment=Alignment(text_rotation=90, horizontal="center") #45)
        
        # if not cell_widths:
        #     # cell_widths=[3,3,10,3,10,25,25,20,3,8,6] + [3]*14 +[40]
        #     # cell_widths=[3,3,10,3,10,12,12,12,3,8,6] + [3]*14 +[40]
        #     cell_widths=[3,3,10,3,10,12,12,12,3,8,6] + [3]*18 +[40]
        # for i, width in enumerate(cell_widths):
        #     ws.column_dimensions[get_column_letter(i+1)].width=width     
        
        # border = Border(left=Side(style='medium'), 
        #              right=Side(style='medium'), 
        #              top=Side(style='medium'), 
        #              bottom=Side(style='medium'))

        # grey_fill= PatternFill(patternType='solid', fgColor=Color('D9D9D9'))

        # for row in ws.iter_rows():
        #     for cell in row:
        #         cell.border = border  
        #         if cell.column_letter in ["C","E","F","G","H","Z"]: # deliberate no wrap on 4 as rare entries and useuall Sheffield and makes row too wide
        #             cell.alignment=cell.alignment.copy(wrap_text=True)
        #         if cell.column_letter in ["A","C","E","G","I","K","M","O","Q","S","U","W","Y"]: # deliberate no wrap on 4 as rare entries and useuall Sheffield and makes row too wide
        #             cell.fill=grey_fill